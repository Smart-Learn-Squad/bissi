"""Excel operations for BISSI.

Provides functionality to read, write, and manipulate XLSX files.
Includes StyledExcel for professional spreadsheets with styled headers and cells.
"""
import openpyxl
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional, Union
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation


# Predefined color palettes for Excel styling
class ExcelColors:
    """Predefined color palette for Excel styling."""
    PRIMARY = "2E86AB"       # Deep blue
    SECONDARY = "212121"      # Dark gray
    ACCENT = "E74C3C"       # Coral red
    SUCCESS = "27AE60"      # Green
    WARNING = "F1C40F"      # Yellow
    INFO = "3498DB"          # Light blue
    LIGHT_GRAY = "95A5A6"
    HEADER_BG = "2E86AB"     # Header background
    HEADER_FONT = "FFFFFF"   # Header font
    ALT_ROW = "F0F8FF"      # Alternating row


class NumberFormats:
    """Excel number format presets."""
    CURRENCY = '#,##0.00 €'
    DECIMAL = '#,##0.00'
    INTEGER = '#,##0'
    PERCENT = '0.00%'
    DATE = 'DD/MM/YYYY'
    DATETIME = 'DD/MM/YYYY HH:MM'
    SCIENTIFIC = '0.00E+00'


def read_excel(file_path: Union[str, Path], 
               sheet_name: Optional[str] = None,
               header: int = 0) -> pd.DataFrame:
    """Read Excel file into pandas DataFrame.
    
    Args:
        file_path: Path to the .xlsx file
        sheet_name: Name of sheet to read, or None for first sheet
        header: Row to use as column headers
        
    Returns:
        DataFrame with Excel data
    """
    if sheet_name:
        return pd.read_excel(file_path, sheet_name=sheet_name, header=header)
    return pd.read_excel(file_path, header=header)


def get_sheet_names(file_path: Union[str, Path]) -> List[str]:
    """Get list of all sheet names in Excel file.
    
    Args:
        file_path: Path to the .xlsx file
        
    Returns:
        List of sheet names
    """
    workbook = openpyxl.load_workbook(file_path, read_only=True)
    sheets = workbook.sheetnames
    workbook.close()
    return sheets


def get_sheet_info(file_path: Union[str, Path]) -> Dict[str, Any]:
    """Get metadata about all sheets in Excel file.
    
    Args:
        file_path: Path to the .xlsx file
        
    Returns:
        Dictionary with sheet names and row counts
    """
    workbook = openpyxl.load_workbook(file_path)
    info = {
        'sheets': {},
        'total_sheets': len(workbook.sheetnames)
    }
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        info['sheets'][sheet_name] = {
            'rows': sheet.max_row,
            'columns': sheet.max_column
        }
    
    workbook.close()
    return info


def write_excel(file_path: Union[str, Path],
                data: Union[pd.DataFrame, List[List[Any]], Dict[str, List[List[Any]]]],
                sheet_name: str = "Sheet1",
                index: bool = False) -> None:
    """Write data to Excel file.
    
    Args:
        file_path: Path for output .xlsx
        data: DataFrame, 2D list, or dict mapping sheet names to 2D lists
        sheet_name: Sheet name for single sheet output
        index: Whether to include DataFrame index
    """
    if isinstance(data, pd.DataFrame):
        data.to_excel(file_path, sheet_name=sheet_name, index=index)
    elif isinstance(data, dict):
        # Multiple sheets
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for sheet, sheet_data in data.items():
                if isinstance(sheet_data, pd.DataFrame):
                    sheet_data.to_excel(writer, sheet_name=sheet, index=index)
                else:
                    pd.DataFrame(sheet_data).to_excel(writer, sheet_name=sheet, index=index)
    else:
        # Single 2D list
        pd.DataFrame(data).to_excel(file_path, sheet_name=sheet_name, index=index)


def add_formula(file_path: Union[str, Path],
                cell: str,
                formula: str,
                sheet_name: Optional[str] = None) -> None:
    """Add formula to specific cell in Excel file.
    
    Args:
        file_path: Path to the .xlsx file
        cell: Cell reference (e.g., 'B5')
        formula: Excel formula string
        sheet_name: Target sheet, or active sheet if None
    """
    workbook = openpyxl.load_workbook(file_path)
    
    if sheet_name and sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active
    
    sheet[cell] = formula
    workbook.save(file_path)
    workbook.close()


def create_chart(file_path: Union[str, Path],
                 chart_type: str,
                 data_range: str,
                 title: str,
                 sheet_name: Optional[str] = None,
                 position: str = "E2") -> None:
    """Create chart in Excel file.
    
    Args:
        file_path: Path to the .xlsx file
        chart_type: 'bar', 'line', or 'pie'
        data_range: Cell range for data (e.g., 'A1:B10')
        title: Chart title
        sheet_name: Target sheet, or active sheet if None
        position: Cell position for chart
    """
    workbook = openpyxl.load_workbook(file_path)
    
    if sheet_name and sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active
    
    # Create chart
    chart_types = {
        'bar': BarChart,
        'line': LineChart,
        'pie': PieChart
    }
    
    chart_class = chart_types.get(chart_type.lower(), BarChart)
    chart = chart_class()
    chart.title = title
    
    # Parse data range and add data
    # Simple range parsing for now
    data = Reference(sheet, min_col=1, min_row=1, max_col=2, max_row=10)
    cats = Reference(sheet, min_col=1, min_row=2, max_row=10)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    sheet.add_chart(chart, position)
    workbook.save(file_path)
    workbook.close()


def get_formulas(file_path: Union[str, Path], 
                 sheet_name: Optional[str] = None) -> Dict[str, str]:
    """Extract all formulas from Excel sheet.
    
    Args:
        file_path: Path to the .xlsx file
        sheet_name: Target sheet, or active sheet if None
        
    Returns:
        Dictionary mapping cell references to formulas
    """
    workbook = openpyxl.load_workbook(file_path, data_only=False)
    
    if sheet_name and sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active
    
    formulas = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).startswith('='):
                formulas[cell.coordinate] = cell.value
    
    workbook.close()
    return formulas


def to_dataframe(file_path: Union[str, Path],
                 sheet_name: Optional[str] = None) -> pd.DataFrame:
    """Convert Excel sheet to pandas DataFrame.
    
    Args:
        file_path: Path to the .xlsx file
        sheet_name: Target sheet, or first sheet if None
        
    Returns:
        DataFrame with sheet data
    """
    return read_excel(file_path, sheet_name=sheet_name)


def summarize_sheet(file_path: Union[str, Path],
                    sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """Generate summary statistics for Excel sheet.
    
    Args:
        file_path: Path to the .xlsx file
        sheet_name: Target sheet, or first sheet if None
        
    Returns:
        Dictionary with summary statistics
    """
    df = read_excel(file_path, sheet_name=sheet_name)
    
    summary = {
        'shape': df.shape,
        'columns': list(df.columns),
        'dtypes': df.dtypes.to_dict(),
        'numeric_summary': df.describe().to_dict() if df.select_dtypes(include=['number']).shape[1] > 0 else {},
        'missing_values': df.isnull().sum().to_dict()
    }
    
    return summary


class StyledExcel:
    """Create polished, styled Excel workbooks with headers, formats, and conditional styling.

    Provides a fluent API for professional spreadsheet generation with:
    - Styled header rows (bold, colored backgrounds)
    - Column width management
    - Frozen panes and filters
    - Number formatting (currency, dates, percentages)
    - Conditional formatting
    - Merged cells
    - Cell borders and alignment

    Example:
        wb = StyledExcel("sales.xlsx")
        wb.add_sheet("Q1 2024")
        wb.add_header_row(["Product", "Qty", "Price", "Total"])
        wb.add_row(["Widget", 100, 9.99], bold=True)
        wb.add_number_format("CURRENCY")
        wb.freeze_panes("B2")
        wb.auto_filter()
        wb.save()
    """

    def __init__(self, file_path: Optional[str] = None):
        """Initialize a styled workbook.

        Args:
            file_path: Path to save the workbook (optional, can set later)
        """
        self.file_path = file_path
        self.workbook = openpyxl.Workbook()
        self.current_sheet = self.workbook.active
        self._column_configs = {}

    @property
    def sheet(self):
        """Get current worksheet."""
        return self.current_sheet

    def add_sheet(self, name: str) -> 'StyledExcel':
        """Add a new worksheet.

        Args:
            name: Sheet name

        Returns:
            self for fluent chaining
        """
        self.current_sheet = self.workbook.create_sheet(name)
        self._column_configs = {}
        return self

    def set_active_sheet(self, name: str) -> 'StyledExcel':
        """Set active sheet by name.

        Args:
            name: Sheet name

        Returns:
            self for fluent chaining
        """
        if name in self.workbook.sheetnames:
            self.current_sheet = self.workbook[name]
            self._column_configs = {}
        return self

    # ─────────────────────────────────────────────────────────
    # Headers
    # ─────────────────────────────────────────────────────────

    def add_header_row(
        self,
        headers: List[str],
        bold: bool = True,
        bg_color: str = None,
        font_color: str = None,
        font_size: int = 11,
        alignment: str = 'center',
    ) -> 'StyledExcel':
        """Add a styled header row.

        Args:
            headers: List of column headers
            bold: Make headers bold
            bg_color: Header background color (hex, default: PRIMARY)
            font_color: Header font color (hex, default: WHITE)
            font_size: Font size
            alignment: 'left', 'center', 'right'

        Returns:
            self for fluent chaining
        """
        bg = bg_color or ExcelColors.HEADER_BG
        font = font_color or ExcelColors.HEADER_FONT

        fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
        font_style = Font(bold=bold, size=font_size, color=font)

        align_map = {'left': 'left', 'center': 'center', 'right': 'right'}
        alignment_style = Alignment(horizontal=align_map.get(alignment, 'center'), vertical='center')

        for col, header in enumerate(headers, start=1):
            cell = self.current_sheet.cell(row=1, column=col)
            cell.value = header
            cell.fill = fill
            cell.font = font_style
            cell.alignment = alignment_style

        return self

    def add_subheader_row(
        self,
        headers: List[str],
        bold: bool = True,
        font_size: int = 10,
    ) -> 'StyledExcel':
        """Add a secondary header row (below main header).

        Args:
            headers: List of sub-headers
            bold: Make bold
            font_size: Font size

        Returns:
            self for fluent chaining
        """
        row = self.current_sheet.max_row + 1
        font_style = Font(bold=bold, size=font_size)

        for col, header in enumerate(headers, start=1):
            cell = self.current_sheet.cell(row=row, column=col)
            cell.value = header
            cell.font = font_style

        return self

    # ─────────────────────────────────────────────────────────
    # Data Rows
    # ─────────────────────────────────────────────────────────

    def add_row(
        self,
        values: List[Any],
        bold: bool = False,
        italic: bool = False,
        font_size: int = 11,
        bg_color: str = None,
    ) -> 'StyledExcel':
        """Add a data row.

        Args:
            values: List of cell values
            bold: Make row bold
            italic: Make row italic
            font_size: Font size
            bg_color: Background color (for entire row)

        Returns:
            self for fluent chaining
        """
        row = self.current_sheet.max_row + 1
        font_style = Font(bold=bold, italic=italic, size=font_size)

        fill = None
        if bg_color:
            fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

        for col, value in enumerate(values, start=1):
            cell = self.current_sheet.cell(row=row, column=col)
            cell.value = value
            cell.font = font_style
            if fill:
                cell.fill = fill

        return self

    def add_rows(
        self,
        data: List[List[Any]],
        bold_rows: List[int] = None,
    ) -> 'StyledExcel':
        """Add multiple data rows.

        Args:
            data: 2D list of values
            bold_rows: Row indices (0-based) to make bold

        Returns:
            self for fluent chaining
        """
        bold_set = set(bold_rows or [])
        for i, row in enumerate(data):
            self.add_row(row, bold=(i in bold_set))
        return self

    def add_data_table(
        self,
        data: List[dict],
        columns: List[str] = None,
        header_row: bool = True,
    ) -> 'StyledExcel':
        """Add table from list of dictionaries.

        Args:
            data: [{col1: val, col2: val}, ...]
            columns: Column order (uses dict keys if None)
            header_row: Include header row

        Returns:
            self for fluent chaining
        """
        if not data:
            return self

        cols = columns or list(data[0].keys())

        # Header
        if header_row:
            self.add_header_row(cols)

        # Data rows
        for row in data:
            self.add_row([row.get(col) for col in cols])

        return self

    # ─────────────────────────────────────────────────────────
    # Number Formats
    # ──────────��──────────────────────────────────────────────

    def set_column_format(
        self,
        column: int,
        format: str,
    ) -> 'StyledExcel':
        """Set number format for a column.

        Args:
            column: Column index (1-based)
            format: Format string or key from NumberFormats

        Returns:
            self for fluent chaining
        """
        format_str = getattr(NumberFormats, format, format) if format.isupper() else format
        max_row = self.current_sheet.max_row

        for row in range(1, max_row + 1):
            cell = self.current_sheet.cell(row=row, column=column)
            cell.number_format = format_str

        return self

    def set_currency(self, column: int) -> 'StyledExcel':
        """Format column as currency."""
        return self.set_column_format(column, NumberFormats.CURRENCY)

    def set_decimal(self, column: int, decimals: int = 2) -> 'StyledExcel':
        """Format column as decimal."""
        format_str = f'#,##{"0" * decimals}.{"0" * decimals}'
        for row in range(1, self.current_sheet.max_row + 1):
            self.current_sheet.cell(row=row, column=column).number_format = format_str
        return self

    def set_percent(self, column: int) -> 'StyledExcel':
        """Format column as percentage."""
        return self.set_column_format(column, NumberFormats.PERCENT)

    def set_date(self, column: int) -> 'StyledExcel':
        """Format column as date."""
        return self.set_column_format(column, NumberFormats.DATE)

    # ─────────────────────────────────────────────────────────
    # Column Configuration
    # ─────────────────────────────────────────────────────────

    def set_column_width(self, column: int, width: float) -> 'StyledExcel':
        """Set column width.

        Args:
            column: Column index (1-based)
            width: Width in characters

        Returns:
            self for fluent chaining
        """
        col_letter = get_column_letter(column)
        self.current_sheet.column_dimensions[col_letter].width = width
        return self

    def set_auto_column_width(self, column: int) -> 'StyledExcel':
        """Auto-fit column width.

        Args:
            column: Column index (1-based)

        Returns:
            self for fluent chaining
        """
        col_letter = get_column_letter(column)
        self.current_sheet.column_dimensions[col_letter].auto_size = True
        return self

    def set_column_widths(self, widths: List[float]) -> 'StyledExcel':
        """Set multiple column widths.

        Args:
            widths: List of widths by column index

        Returns:
            self for fluent chaining
        """
        for i, width in enumerate(widths, start=1):
            self.set_column_width(i, width)
        return self

    # ─────────────────────────────────────────────────────────
    # Freeze & Filter
    # ─────────────────────────────────────────────────────────

    def freeze_panes(self, cell: str = 'B2') -> 'StyledExcel':
        """Freeze panes at cell position.

        Args:
            cell: Cell reference (e.g., 'B2' freezes row 1 and column A)

        Returns:
            self for fluent chaining
        """
        self.current_sheet.freeze_panes = cell
        return self

    def freeze_header(self) -> 'StyledExcel':
        """Freeze header row."""
        return self.freeze_panes('A2')

    def auto_filter(self, ref: str = None) -> 'StyledExcel':
        """Add auto-filter to range.

        Args:
            ref: Range reference (e.g., 'A1:E10'). Uses all data if None.

        Returns:
            self for fluent chaining
        """
        if ref:
            self.current_sheet.auto_filter.ref = ref
        else:
            # Auto-filter all data
            max_row = self.current_sheet.max_row
            max_col = self.current_sheet.max_column
            if max_row > 0 and max_col > 0:
                ref = f'A1:{get_column_letter(max_col)}{max_row}'
                self.current_sheet.auto_filter.ref = ref
        return self

    # ─────────────────────────────────────────────────────────
    # Cell Styling
    # ─────────────────────────────────────────────────────────

    def set_cell_color(
        self,
        row: int,
        column: int,
        bg_color: str,
        font_color: str = None,
    ) -> 'StyledExcel':
        """Set cell background color.

        Args:
            row: Row index (1-based)
            column: Column index (1-based)
            bg_color: Background color (hex)
            font_color: Font color (hex, optional)

        Returns:
            self for fluent chaining
        """
        cell = self.current_sheet.cell(row=row, column=column)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        if font_color:
            cell.font = Font(color=font_color)
        return self

    def apply_borders(
        self,
        start_row: int = 1,
        end_row: int = None,
        start_col: int = 1,
        end_col: int = None,
        style: str = 'thin',
    ) -> 'StyledExcel':
        """Apply borders to range.

        Args:
            start_row: Start row (1-based)
            end_row: End row (defaults to max row)
            start_col: Start column (1-based)
            end_col: End column (defaults to max col)
            style: 'thin', 'medium', 'thick'

        Returns:
            self for fluent chaining
        """
        border_styles = {
            'thin': Side(style='thin'),
            'medium': Side(style='medium'),
            'thick': Side(style='thick'),
        }
        side = border_styles.get(style, Side(style='thin'))
        border = Border(left=side, right=side, top=side, bottom=side)

        end_row = end_row or self.current_sheet.max_row
        end_col = end_col or self.current_sheet.max_column

        for row_idx in range(start_row, end_row + 1):
            for col_idx in range(start_col, end_col + 1):
                cell = self.current_sheet.cell(row=row_idx, column=col_idx)
                cell.border = border

        return self

    def set_cell_alignment(
        self,
        row: int,
        column: int,
        horizontal: str = 'left',
        vertical: str = 'center',
    ) -> 'StyledExcel':
        """Set cell alignment.

        Args:
            row: Row index (1-based)
            column: Column index (1-based)
            horizontal: 'left', 'center', 'right'
            vertical: 'top', 'center', 'bottom'

        Returns:
            self for fluent chaining
        """
        cell = self.current_sheet.cell(row=row, column=column)
        cell.alignment = Alignment(horizontal=horizontal, vertical=vertical)
        return self

    # ─────────────────────────────────────────────────────────
    # Merged Cells
    # ─────────────────────────────────────────────────────────

    def merge_cells(
        self,
        start_row: int,
        end_row: int,
        start_col: int,
        end_col: int,
    ) -> 'StyledExcel':
        """Merge cells.

        Args:
            start_row: Start row (1-based)
            end_row: End row (1-based)
            start_col: Start column (1-based)
            end_col: End column (1-based)

        Returns:
            self for fluent chaining
        """
        start = f'{get_column_letter(start_col)}{start_row}'
        end = f'{get_column_letter(end_col)}{end_row}'
        self.current_sheet.merge_cells(f'{start}:{end}')
        return self

    # ─────────────────────────────────────────────────────────
    # Conditional Formatting
    # ─────────────────────────────────────────────────────────

    def add_color_scale(
        self,
        start_row: int,
        end_row: int,
        start_col: int,
        end_col: int,
    ) -> 'StyledExcel':
        """Apply color scale (gradient from low to high values).

        Args:
            start_row: Start row
            end_row: End row
            start_col: Start column
            end_col: End column

        Returns:
            self for fluent chaining
        """
        ref = f'{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}'
        rule = ColorScaleRule(
            start_type='min', start_color=ExcelColors.INFO,
            mid_type='percentile', mid_value=50, mid_color='FFF2CC',
            end_type='max', end_color=ExcelColors.ACCENT,
        )
        self.current_sheet.conditional_formatting.add(ref, rule)
        return self

    # ─────────────────────────────────────────────────────────
    # Data Validation
    # ─────────────────────────────────────────────────────────

    def add_dropdown(
        self,
        row: int,
        column: int,
        options: List[str],
    ) -> 'StyledExcel':
        """Add dropdown validation to cell.

        Args:
            row: Row index (1-based)
            column: Column index (1-based)
            options: List of dropdown options

        Returns:
            self for fluent chaining
        """
        dv = DataValidation(type='list', formula1=f'"{",".join(options)}"')
        self.workbook.add_data_validation(dv)
        cell = f'{get_column_letter(column)}{row}'
        dv.add(self.current_sheet[cell])
        return self

    # ─────────────────────────────────────────────────────────
    # Save
    # ─────────────────────────────────────────────────────────

    def save(self, file_path: str = None) -> None:
        """Save the workbook.

        Args:
            file_path: Path to save (uses init path if None)
        """
        path = file_path or self.file_path
        if not path:
            raise ValueError("No file_path specified")
        self.workbook.save(path)

    def get_workbook(self) -> openpyxl.Workbook:
        """Get underlying openpyxl Workbook."""
        return self.workbook
